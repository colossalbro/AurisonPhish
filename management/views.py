from django.utils.decorators import method_decorator
from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.views import View

from .models import Campaign, AurisonUser
from .decorators import lazyAuthorization
from mails.task import sendEmails
from .task import deleteFile
from .utils import *

from os import getcwd, path as osPath
from datetime import date, datetime
from openpyxl import load_workbook
from random import choice
from time import time




class Login(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'login.html')
    
    def post(self, request, *args, **kwargs):
        password = request.POST.get('password')
        
        if password.startswith('>>') and password.endswith('<<'):
            if len(password) > 4:            
                response = redirect('/eganam/camp')    
                response.set_cookie('manage_cookie', choice(sessions), expires=3600, path='/')  # note (to future self) 'sessions' is from .utils 
                return response
            

        return alertRedirect(request, 'Invalid username or password', 'back')





@method_decorator(lazyAuthorization, name='dispatch')
class Logout(View):
    def get(self, request, *args, **kwargs):
        response = redirect('/')
        response.set_cookie('manage_cookie', expires=0)
        return response
    




@method_decorator(lazyAuthorization, name='dispatch')
class Home(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'manage.html')






@method_decorator(lazyAuthorization, name='dispatch')
class NewCampaign(View):
    def get(self, request, *args, **kwargs):
        today = date.today()
        print(today)
        return render(request, 'new_camp.html', {'today': today})
    


    #Trying to keep things simple and avoid using DRF. So, just did things manually. So much for simplicity :(
    #Note (to self): Probably should use DRF to clean this method up.
    def post(self, request, *args, **kwargs):
        try:
            campName  = request.POST.get('name')        #Grab Campaign name.
            end = request.POST.get('date')

            endDate = datetime.strptime(end, '%Y-%m-%d')
            validDate = self.validateDate(request, endDate)

            if not validDate['valid']:
                return validDate['response']
            

            # today = timezone.now().date
            # if endDate.date == today:
            #     return alertRedirect(request, 'campaign cannot end the same day it\'s created', '/eganam/camp/new')
            
            # if endDate.date < today:
            #     return alertRedirect(request, 'Date cannot be in the past', '/eganam/camp/new')

            upload = f'{ int(time()) }__{ request.FILES["file"] }'  #create unique name for the upload.

            uploadFilePath = osPath.join( getcwd(), 'management', 'uploads', upload)    #upload path

            successful = self.saveFile(request.FILES['file'], uploadFilePath)
            
            if successful:
                campaign = Campaign(name=campName, endDate = endDate)
                campaign.save()

                users = self.parseFile(uploadFilePath, campaign)

                deleteFile.delay(uploadFilePath)  #clean up a processed file.
                return redirect(f'/eganam/camp/{campaign.pk}')

            return alertRedirect(request, 'Error parsing file', 'back')
                
        except Exception as e:
            print(e)
            return alertRedirect(request, 'An unknown error has occured', 'back')



    def saveFile(self, xlsxFile, uploadPath):
        try:
            with open(uploadPath, 'wb') as dest:
                for chunk in xlsxFile.chunks():
                    dest.write(chunk)
                
                return True
        except:
            return False


    
    def parseFile(self, path: str, camp: Campaign) -> list:
        try:
            users = []
            book = load_workbook(path, read_only=True)
            sheet = book.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    user = AurisonUser(
                        title = row[0], 
                        firstName = row[1],
                        lastName = row[2],
                        email = row[3],
                        organization = row[4],
                        campaign = camp
                    )
                    user.save()
                    users.append(user)

                except Exception as e:
                    print(f"Error parsing row: {e}")

            return users    #Not even sure why I even return this :D
        
        except Exception as e:
            #Likely an invalid file path:
            print(f"Error parsing file: {e}")
            return []
            


    def validateDate(self, request: HttpRequest, dateObj: datetime) -> dict:
        today = timezone.now().date()

        if dateObj.date() == today:
            return {
                'valid' : False, 
                'response' : alertRedirect(request, 'End date cannot be the same as the creation day', 'back')
            }
        
        if dateObj.date() < today:
            return {
                'valid' : False,
                'response' : alertRedirect(request, 'End date cannot be in the past', 'back')
            }
        
        return {'valid' : True}





# @method_decorator(lazyAuthorization, name='dispatch')
class GetCampaign(View):
    def get(self, request, *args, **kwargs):
        if not kwargs.get('id'):
            if request.GET.get('all').lower() == 'true':
                return self.getAllCampaigns(request)
            
            return redirect('/eganam/camp')


        id = kwargs.get('id')
        try:
            campaign = Campaign.objects.get(pk = id)
            users = AurisonUser.objects.filter(campaign = campaign)
        except:
            return alertRedirect(request, 'campaign not found', '/eganam/camp')
        
        return render(request, 'camp.html', {'campaign' : campaign, 'users' : users})
    


    def getAllCampaigns(self, request):
        campaigns = Campaign.objects.all()

        if len(campaigns) < 1:
            return alertRedirect(request, 'No campaigns found!', '/eganam/camp')
        
        return render(request, 'all_camps.html', {'campaigns' : campaigns})





@method_decorator(lazyAuthorization, name='dispatch')
class startCampaign(View):
    def get(self, request, *args, **kwargs):
        campaign = Campaign.objects.get(pk = kwargs['id'])
       

        if campaign.closed:
            return alertRedirect(request, 'Cannot activate closed campaign', 'back')
        
        if campaign.active:
            return alertRedirect(request, 'Campaign already active!', 'back')
        
        if not campaign.sentEmails:
            sendEmails.delay(campaign.pk)
            return alertRedirect(request, 'Campaign started! Emails in queue', f'/eganam/camp/{kwargs["id"]}')

        
        campaign.active = True
        campaign.save()

        return alertRedirect(request, 'Campaign Activated!', 'back')
    
        



@method_decorator(lazyAuthorization, name='dispatch')
class stopCampaign(View):
    def get(self, request, *args, **kwargs):
        campaign = Campaign.objects.get(pk = kwargs['id'])

        if not campaign.active:
            return alertRedirect(request, 'Campaign not active!', 'back')
        
        campaign.active = False
        campaign.save()

        return alertRedirect(request, 'Campaign Deactivated! Phishing attempts redirect to aurison.app', 'back')





@method_decorator(lazyAuthorization, name='dispatch')
class DeleteCampaign(View):
    def get(self, request, *args, **kwargs):
        try:
            campagin = Campaign.objects.get(pk=kwargs['id'])

            campagin.delete()

            return alertRedirect(request, 'Campaign deleted!', '/eganam/camp')
            
        except:
            return alertRedirect(request, 'Campaign not found', '/eganam/camp/all')
        




@method_decorator(lazyAuthorization, name='dispatch')
class CampaignStats(View):
    def get(self, request, *args, **kwargs):
        try:
            campaign = Campaign.objects.get(pk = kwargs['id'])
        except:
            #Campaign likely doesn't exist
            return alertRedirect(request, 'Invalid Campaign', '/eganam/camp')


        allUsers = AurisonUser.objects.filter(campaign = campaign)

        #Users who submitted their passwords
        phished = AurisonUser.objects.filter(campaign = campaign, submittedPass = True)

        #Users who clicked the link but didn't submit their passwords.
        clicked = AurisonUser.objects.filter(campaign = campaign, clickedLink = True, submittedPass = False)

        #Users who opened the email and ignored it.
        ignored = AurisonUser.objects.filter(campaign = campaign, openedEmail = True, clickedLink = False, submittedPass = False)

        fileName = f'{ int(time()) }-{campaign.name}.xlsx'  #create unique name for the upload.
        self.excelFile = osPath.join( getcwd(), 'management', 'static', 'files', fileName)  #path to excel file


        try:

            genCampaignAnalytics(self.excelFile, allUsers, clicked, phished, ignored)

        except Exception as e:
            print(e)        #Again need to setup logging or something.
            return HttpResponse(status=500)


        return FileResponse(open(self.excelFile, 'rb'))