from django.http import HttpResponse, HttpRequest
from django.shortcuts import render
from django.utils import timezone
from .models import AurisonUser
from .task import deleteFile


from datetime import timedelta
from openpyxl import Workbook





def alertRedirect(request: HttpRequest, message: str, href: str) -> HttpResponse:
    return render(
        request,
        'redirect.html',
        {'message' : message, 'href' : href}
    )






def genCampaignAnalytics(excelPath: str, users: list[AurisonUser], clicked: list[AurisonUser], phished: list[AurisonUser], ignored: list[AurisonUser]):
    workbook = Workbook()      
    analytics = workbook.active
    analytics.title = 'Analytics'


    percentClick = f'{len(clicked)} of {len(users)} users clicked the phishing link. { ( len(clicked) / len(users) ) * 100:.2f}%'
    percentPhish = f'{len(phished)} of {len(users)} users submitted their passwords. { ( len(phished) / len(users) ) * 100:.2f}%'
    percentIgnore = f'{len(ignored)} of {len(users)} users ignored the email. { ( len(ignored) / len(users) ) * 100:.2f}%'


    analytics.cell(row=1, column=1, value=percentIgnore)
    analytics.cell(row=2, column=1, value=percentClick)
    analytics.cell(row=3, column=1, value=percentPhish)


    #Other entries.
    excelData = {
        'IGNORED' : [[user.title, user.firstName, user.lastName, user.email, user.organization] for user in phished],
        'CLICKED' : [[user.title, user.firstName, user.lastName, user.email, user.organization] for user in clicked],
        'PHISHED' : [[user.title, user.firstName, user.lastName, user.email, user.organization] for user in ignored]
    }

    for name, entries in excelData.items():
        sheet = workbook.create_sheet(title=name)
        header = ['Title', 'First Name', 'Last Name', 'Email', 'Organization']

        data = [header]         #This way the headers come first when the entries are being written.
        data.extend(entries)    #Everything else :)

        #Write data to sheet.
        for rowNum, rowData in enumerate(data, start=1):                #I do this to get an index (rowNum)
            for colNum, value in enumerate(rowData, start=1):           #Same, need to get an index (colNum)
                sheet.cell(row=rowNum, column=colNum, value=value)      #write data into cell


    workbook.save(excelPath)    #Create the excel File 
    

    #Delete the excel file 5 minutes after it's created.
    #Should be enough time to return to the frontend or email it.
    eta = timezone.now() + timedelta(seconds=300)
    deleteFile.apply_async(args=[excelPath], eta=eta)






sessions = [
    "kjsnsfjnsfsfjnsjlfnmsjnfsijwnemoki2n24irwjhd9hn30jf2n4",
    "927h4un2mr9-hb2d1e0-rjh2bw80urtyh89rn3oi2r9m03wfpk0k4o",
    "me2irh98th9eu1j0m-e9jn0i2edfe1-jf9rm4nwduh2n49rmnfiin0",
    "092i89hgyrbn1-0oifjvn1ev9rj0nimr21e-r9jnmges-j0atim4-f",
    "-e1908rhn2mjri2n3owdpm902mi0htjmopaioni9jfm2oirwfnem02",
    ")*(fgyb2ner02fugubanr29km[paz,-rfnoegaubnfmp2-9jg0nf20",
    "FMnof-1ikmfn2=]Md2k,1[d=mjqkmnfij3akmopi2j=fm--v0h9jm1",
    "1ijnf-0kmpniogaj-kmospanrmvfm=w0ajnzdj2vm2p-krmpjmaknf",
    "fisoj2m-03jem10-j2pon3vjkampiobgnj-raopemfnpjopkmrjgks",
    "opj-kmp2nrjvmrij-vmpkijo2nfjognao04irjiogjanmfwpiwkmzm",
    "-90u81h9runf-a9jnifrgeuhajrm2oinfbgazjf0wmu290hinofk0e",
    "0iekjrfniogajdlszdmvon=0-jrnigdadzlngoi2klrnfoigjaniml",
    "pimwo390894hn2oe0-ikjr2nibaeonrnm2r9280hnrn19j03nrjfms",
    "d208hfnoae9-mi2rubfgaj90im1erg-jfmop2rnughjpmid2kfnjom",
    "j0dhu9bf3u9b1h0-jn2fi3ghr29j0tnoigibsnarm2inoughjrm2pr",
    ")9d728f6g7h8j0aimnyfdg9h1j-2fnio2-r0h93ghrjke1rmionigb",
    "-908h29grhjmdfughf8dj91kjmfniubfh8j-1mnbuih2fjkmfo2gnf",
    "mnufbygf2hjkf0j9vuy8ghj1km2v3nijkc,mvnshhajr2ngbjfm2gn",
    "kjfu32-i=kjgim3nbujk2v93j0ibnsuomvf3kof09j20i3gnmfkvmr",
    "KJfn2u-3jvkmpoinou2hf3vrk-0moi2onfgjjk,p3gnonjgk-2o,fm",
    "10i2-9r08h9gjaj-30nteaf0j-m0ghrtj-0in2r3g=0k-m0iauaimf",
    "1-0ir290htugfijdmO1ejritnumf2dofi0oubvhv2-jni0vuobsnwz",
    "0d-j92h0fubi3yvvnjk-0avsmpioundzmvpijaovsnoamfjfagmsie",
    "-2r90893hfjmisvounapkmwofnjigsjkfmkjvoihbvj2mp3inubhj2",
    "19j02fngmapcovsjbnjamif2ouhv3j-kcmp2cinovubsmzWPRonagj",
    "0=eirjignmapifojkmvs0iorbuhvj=2q-vi3nbbh9b8jrm2tg0hjmf",
    "-012jf0ghjfkm2i3vouh1cj-2mi3n4urbjit0noghijsbijifnmpvs",
    "cihn2uobn3v=-9h042rtngimeavpdzinvoo-vjibnsr2j=0-3omgir",
    "10ek-jr2ifngoifj02c-k3vsim02qu3hvj-2mo3pibnujgmpironow",
    "if2nbuv3h1c-i2onugbskljoalm.nn-rpifgpsiz,a-jrintgo2r2g",
    "fbsndmap2h08bun1-r9j3htn1em-r9jhnfdm9fjn1mdkjnfuhni.kd",
    "10-9e08r97gfhjd1k0-jfhuvygag8jeIjidhufhjviembuis9u.jmo",
    "OCIYUVTU9H180Jsnjbkdhugbnqmwc9uy8ghmcjhbeih-786aacmisc"
]